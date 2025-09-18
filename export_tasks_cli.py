import os
import sys
import argparse
from getpass import getpass
from typing import Any, Dict, List, Optional, Tuple

from iquavis_client import IQuavisClient
from excel_writer import collect_headers, flatten_dict, write_tasks_xlsx


DEFAULT_INCLUDES = [
    # Include commonly useful nested resources per samples
    "Assigns",
    "TaskBarStyle",
    "Todos",
    "OutputPlanDeliverables",
    "Progress",
]

DEFAULT_EXTRA_HEADERS = [
    # Canonical Task properties (derived from ref/iquavis.py param_example)
    # Core
    "Id", "Name", "Type", "StartDate", "EndDate", "ProjectId", "TaskDomainId",
    # Calendar
    "Calendar.Sunday", "Calendar.Monday", "Calendar.Tuesday", "Calendar.Wednesday",
    "Calendar.Thursday", "Calendar.Friday", "Calendar.Saturday", "Calendar.Days",
    # Links/Classification
    "ParentTaskId", "ContentsType", "PeriodSummaryType", "ClassId", "TagIds",
    # Meta & URLs
    "AssignCategory", "Priority", "UrlTitle", "Url",
    # Selections
    "TaskManualSelection", "TaskAutoSelection",
    # Custom fields
    "CustomAttributes",
    # Man hour
    "ManHourUnit", "TemplateStandardManHour", "IndependentStandardManHour", "EnableOperatingHourRate",
    # Assignments & progress
    "Assigns", "ActualManHourType", "SumUpProject", "SumUpGroupTask", "ProgressType",
    # Alerts
    "EnableCompleteAlert", "CompleteAlertCautionDelayBaseDays", "CompleteAlertDelayBaseRate",
    "CompleteAlertDeadlineType", "CompleteAlertDeadlineDays",
    "EnableOnGoingAlert", "OnGoingAlertCautionDelayBaseDays", "OnGoingAlertWarningDelayBaseDays",
    "OnGoingAlertDelayBaseRate", "EnableDeadlineAlert", "EnableDeadlineAlertBaseDays",
    "DeadlineAlertBaseDays", "EnablePreviousAlert", "EnablePreviousAlertBaseDays",
    "PreviousAlertBaseDays",
    # Other
    "MailType", "Todos", "OutputPlanDeliverables", "Note", "ReferenceCode",
    "DeadlineTaskType", "DeadlineTaskId", "PlannedStartDateFixedFlag", "EarliestStartDate", "LatestEndDate",
    "PlannedDurationOptimistic", "PlannedDurationPessimistic", "InformationSharingTiming", "PaceCreatingInformation",
    "InputRiskConsiderationFlag", "ProficiencyLevel", "ProficiencyLimit", "FixReworkAmountFlag", "SimulationFlag",
    "TaskDomainRowNumber", "AndonId",
    # Task bar style
    "TaskBarStyle.Shape", "TaskBarStyle.Pattern", "TaskBarStyle.Background", "TaskBarStyle.IsTwoRows",
    "TaskBarStyle.KeepStyle", "TaskBarStyle.VisibleEndLine", "TaskBarStyle.IsLayoutCheck", "TaskBarStyle.BarAlignment",
]


def prompt_login(debug: bool = False) -> IQuavisClient:
    base_url = os.getenv("IQUAVIS_BASE_URL")  # Optional override
    client = IQuavisClient(base_url=base_url, debug=debug)

    print("iQUAVIS login")
    user_id = input("User ID: ").strip()
    password = getpass("Password: ")
    try:
        client.login(user_id, password)
        print("Authenticated successfully.")
    except Exception as e:
        print(f"Login failed: {e}")
        sys.exit(1)
    return client


def choose_project(projects: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not projects:
        print("No projects available for this user.")
        sys.exit(1)

    # Print with numbering
    print("\nProjects (select by number):")
    for idx, p in enumerate(projects, start=1):
        pid = str(p.get("Id") or p.get("id") or p.get("ID") or "")
        name = str(p.get("Name") or p.get("name") or "")
        print(f"  {idx:>3}. {name}  [Id: {pid}]")

    while True:
        raw = input(f"Enter number 1-{len(projects)}: ").strip()
        if not raw.isdigit():
            print("Please enter a valid number.")
            continue
        n = int(raw)
        if not (1 <= n <= len(projects)):
            print("Out of range.")
            continue
        return projects[n - 1]


def load_projects_from_sheet(path: str) -> Tuple[Optional[List[List[Any]]], Optional[List[Dict[str, Any]]]]:
    """Load project rows and dictionaries from an existing Excel workbook."""
    if not path:
        return None, None
    if not os.path.exists(path):
        print("Specified Excel file not found; fetching projects from server.")
        return None, None
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"openpyxl is required to load existing workbook: {e}")
        return None, None

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"Failed to load workbook '{path}': {e}")
        return None, None

    if "project" not in wb.sheetnames:
        return None, None
    ws = wb["project"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None, None

    headers = [str(h) if h is not None else "" for h in rows[0]]
    projects: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        proj = {headers[i]: r[i] for i in range(len(headers))}
        projects.append(proj)

    if not projects:
        return None, None
    return rows, projects


def select_existing_excel_file() -> str:
    """Open a GUI dialog to select an existing Excel file.

    Returns the chosen path or an empty string if the user cancels or if
    ``tkinter`` is unavailable.
    """
    try:
        from tkinter import Tk, filedialog
    except Exception as e:
        print(f"tkinter is required for file selection: {e}")
        return ""

    try:
        root = Tk()
        root.withdraw()
        path = filedialog.askopenfilename(
            title="Select existing Excel file (Cancel for none)",
            filetypes=[("Excel files", "*.xlsm *.xlsx *.xls"), ("All files", "*.*")],
        )
        root.destroy()
        return path.strip()
    except Exception as e:
        print(f"File selection dialog failed: {e}")
        return ""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()

    client = prompt_login(debug=args.debug)

    print("Select existing Excel file (Cancel if none).")
    existing_path = select_existing_excel_file()
    project_rows, projects = load_projects_from_sheet(existing_path)

    if not projects:
        try:
            projects = client.list_projects()
        except Exception as e:
            print(f"Failed to fetch projects: {e}")
            sys.exit(1)
        flat_projects = [flatten_dict(p) for p in projects]
        proj_headers = collect_headers(flat_projects)
        project_rows = [proj_headers] + [[fp.get(h) for h in proj_headers] for fp in flat_projects]

    project = choose_project(projects)
    proj_id, proj_name = client.project_identity(project)

    # Fetch tasks for the selected project
    try:
        tasks = client.list_tasks(proj_id, include=DEFAULT_INCLUDES)
    except Exception as e:
        print(f"Task fetch with includes failed, retrying without includes... ({e})")
        try:
            tasks = client.list_tasks(proj_id, include=None)
        except Exception as e2:
            print(f"Failed to fetch tasks: {e2}")
            sys.exit(1)

    print(f"Fetched {len(tasks)} tasks from '{proj_name}'.")

    # Unwrap tasks if response has Task wrapper (defensive)
    tasks_unwrapped: List[Dict[str, Any]] = [client.unwrap_task(t) for t in tasks]

    # Write to Excel in the same directory as this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        out_path = write_tasks_xlsx(
            tasks_unwrapped,
            proj_name,
            script_dir,
            extra_headers=DEFAULT_EXTRA_HEADERS,
            project_sheet_rows=project_rows,
        )
    except Exception as e:
        print(f"Export failed: {e}")
        sys.exit(1)

    print(f"Exported to: {out_path}")


if __name__ == "__main__":
    main()
