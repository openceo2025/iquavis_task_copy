import argparse
import json
import math
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from getpass import getpass
from typing import Any, Dict, Iterable, List, Optional, Tuple

from iquavis_client import IQuavisClient


YELLOW_RGBS = {"FFFFFF00", "FFFF00", "00FFFF00"}
YELLOW_INDEXED = {5, 6, 44}
BLUE_RGB = "FF00B0F0"
RED_RGB = "FFFF0000"


@dataclass
class TaskRow:
    row_index: int
    task_id: str
    project_id: str
    flat_values: Dict[str, Any]
    yellow_cells: List[Tuple[int, int]]  # (row, column)


def prompt_login(debug: bool = False) -> IQuavisClient:
    base_url = os.getenv("IQUAVIS_BASE_URL")
    client = IQuavisClient(base_url=base_url, debug=debug)

    print("iQUAVIS login")
    user_id = input("User ID: ").strip()
    password = getpass("Password: ")
    try:
        client.login(user_id, password)
        print("Authenticated successfully.")
    except Exception as exc:  # pragma: no cover - network failure
        print(f"Login failed: {exc}")
        sys.exit(1)
    return client


def select_excel_file() -> str:
    try:
        from tkinter import Tk, filedialog
    except Exception as exc:
        print(f"tkinter is required for file selection: {exc}")
        return ""

    try:
        root = Tk()
        root.withdraw()
        path = filedialog.askopenfilename(
            title="Select Excel file to import",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        root.destroy()
        return path.strip()
    except Exception as exc:
        print(f"File selection dialog failed: {exc}")
        return ""


def _normalize_numeric(value: float) -> Any:
    if math.isfinite(value) and float(int(value)) == float(value):
        return int(value)
    return value


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float):
        return _normalize_numeric(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        lowered = stripped.lower()
        if lowered in {"true", "false"}:
            return lowered == "true"
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, (dict, list, int, float, bool)):
                return parsed
        except json.JSONDecodeError:
            pass
        return stripped
    return value


def unflatten(flat: Dict[str, Any], sep: str = ".") -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    for key, value in flat.items():
        if value is None:
            continue
        parts = [part for part in key.split(sep) if part]
        if not parts:
            continue
        cursor = result
        for part in parts[:-1]:
            cursor = cursor.setdefault(part, {})
        cursor[parts[-1]] = value
    return result


def _cell_rgb(cell: Any) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if not fill or fill.patternType != "solid":
        return None
    color = fill.start_color
    if color is None:
        return None
    if getattr(color, "type", None) == "rgb":  # type: ignore[attr-defined]
        rgb = (color.rgb or "").upper()
        return rgb
    if getattr(color, "type", None) == "indexed":  # type: ignore[attr-defined]
        idx = getattr(color, "indexed", None)
        if idx is not None and idx in YELLOW_INDEXED:
            return "INDEXED_YELLOW"
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and rgb.upper() in YELLOW_RGBS:
        return rgb.upper()
    return None


def collect_task_rows(ws: Any, original_ws: Optional[Any] = None) -> List[TaskRow]:
    headers: List[str] = []
    rows: List[TaskRow] = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        if row[0].row == 1:
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
            continue
        if not headers:
            break

        values = {}
        yellow_cells: List[Tuple[int, int]] = []
        has_value = False
        for idx, cell in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            cell_value = cell.value
            if cell_value not in (None, ""):
                has_value = True
            rgb = _cell_rgb(cell)
            coord = (cell.row, cell.column)
            if rgb and (rgb in YELLOW_RGBS or rgb == "INDEXED_YELLOW"):
                yellow_cells.append(coord)
            normalized_value = normalize_value(cell_value)
            values[header] = normalized_value

            if original_ws is not None:
                original_cell = original_ws.cell(row=cell.row, column=cell.column)
                original_value = normalize_value(original_cell.value)
                if normalized_value != original_value and coord not in yellow_cells:
                    yellow_cells.append(coord)

        if not has_value:
            continue

        if not yellow_cells:
            continue

        task_id = str(values.get("Id") or values.get("ID") or values.get("TaskId") or "").strip()
        project_id = str(values.get("ProjectId") or values.get("project_id") or "").strip()
        rows.append(
            TaskRow(
                row_index=row[0].row,
                task_id=task_id,
                project_id=project_id,
                flat_values=values,
                yellow_cells=yellow_cells,
            )
        )
    return rows


def set_fill(ws: Any, coords: Iterable[Tuple[int, int]], rgb: str) -> None:
    from openpyxl.styles import PatternFill

    fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
    for row_idx, col_idx in coords:
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill


def save_with_suffix(wb: Any, original_path: str, suffix: str = "_result") -> str:
    directory = os.path.dirname(original_path)
    base = os.path.basename(original_path)
    stem, ext = os.path.splitext(base)
    new_name = f"{stem}{suffix}{ext}"
    out_path = os.path.join(directory, new_name)
    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", action="store_true", help="Enable HTTP debug output")
    args = parser.parse_args()

    client = prompt_login(debug=args.debug)
    print("Select Excel file to import.")
    excel_path = select_excel_file()
    if not excel_path:
        print("No Excel file selected.")
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"openpyxl is required to load Excel files: {exc}")
        sys.exit(1)

    keep_vba = os.path.splitext(excel_path)[1].lower() == ".xlsm"
    try:
        wb = load_workbook(excel_path, keep_vba=keep_vba)
    except Exception as exc:
        print(f"Failed to load workbook: {exc}")
        sys.exit(1)

    tasks_sheet_name = "tasks"
    if tasks_sheet_name not in wb.sheetnames:
        print("The workbook does not contain a 'tasks' sheet.")
        sys.exit(1)

    ws_tasks = wb[tasks_sheet_name]
    original_sheet_name = f"{tasks_sheet_name}_original"
    ws_tasks_original = wb[original_sheet_name] if original_sheet_name in wb.sheetnames else None

    task_rows = collect_task_rows(ws_tasks, original_ws=ws_tasks_original)
    if not task_rows:
        print("No rows with yellow cells were found in the 'tasks' sheet.")
        if ws_tasks_original is None:
            print(
                "If you expected updates to be detected automatically, add a 'tasks_original' "
                "sheet or highlight the cells to import."
            )
        sys.exit(0)

    total = len(task_rows)
    success = 0
    failures: List[Tuple[TaskRow, str]] = []

    for task_row in task_rows:
        if not task_row.task_id:
            failures.append((task_row, "Task Id is missing."))
            set_fill(ws_tasks, task_row.yellow_cells, RED_RGB)
            continue
        if not task_row.project_id:
            failures.append((task_row, "Project Id is missing."))
            set_fill(ws_tasks, task_row.yellow_cells, RED_RGB)
            continue

        payload = unflatten(task_row.flat_values)
        try:
            json.dumps(payload)
        except TypeError:
            failures.append((task_row, "Payload contains non-serializable values."))
            set_fill(ws_tasks, task_row.yellow_cells, RED_RGB)
            continue

        try:
            client.update_task(task_row.project_id, task_row.task_id, payload)
        except Exception as exc:  # pragma: no cover - network failure
            failures.append((task_row, str(exc)))
            set_fill(ws_tasks, task_row.yellow_cells, RED_RGB)
            continue

        success += 1
        set_fill(ws_tasks, task_row.yellow_cells, BLUE_RGB)

    out_path = save_with_suffix(wb, excel_path)

    print("\nImport summary")
    print("-------------")
    print(f"Total rows processed : {total}")
    print(f"Successful updates   : {success}")
    print(f"Failed updates       : {len(failures)}")
    if failures:
        print("\nFailures:")
        for task_row, reason in failures:
            print(f"  Row {task_row.row_index} (Task ID: {task_row.task_id or 'N/A'}): {reason}")
    print(f"\nWorkbook saved to: {out_path}")


if __name__ == "__main__":
    main()
