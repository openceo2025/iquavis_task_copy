import json
import os
import re
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_TEMPLATE_DIR = os.path.join(BASE_DIR, "excel_template")
DEFAULT_TEMPLATE_NAME = "iquavis_tasks.xlsm"


def _is_primitive(x: Any) -> bool:
    return x is None or isinstance(x, (str, int, float, bool))


def flatten_dict(obj: Dict[str, Any], parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
    """
    Flatten nested dicts using dot notation.
    - Lists/tuples are serialized as compact JSON strings to preserve information without column explosion.
    - None stays None; consumers may render as blank.
    """
    items: List[Tuple[str, Any]] = []
    for k, v in (obj or {}).items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else str(k)
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, (list, tuple)):
            try:
                items.append((new_key, json.dumps(v, ensure_ascii=False)))
            except Exception:
                # Fallback to str if not JSON-serializable
                items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)


def collect_headers(rows: Iterable[Dict[str, Any]], extra_headers: Iterable[str] = ()) -> List[str]:
    """
    Build the union of keys from flattened rows. Returns an ordered list:
    - Start with a preferred key order, then append remaining keys sorted.
    """
    preferred = [
        "Id",
        "Name",
        "Type",
        "StartDate",
        "EndDate",
        "ProjectId",
        "TaskDomainId",
    ]
    seen: Set[str] = set(extra_headers or [])
    for row in rows:
        seen.update(row.keys())

    ordered: List[str] = []
    for key in preferred:
        if key in seen:
            ordered.append(key)
            seen.remove(key)
    # Append remaining keys in stable sorted order for predictability
    ordered.extend(sorted(seen))
    return ordered


def sanitize_filename(name: str) -> str:
    # Remove/replace characters invalid on common filesystems
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]", "_", name)  # Windows-forbidden
    # Avoid trailing dots/spaces on Windows
    name = name.rstrip(" .")
    return name or "project"


def next_available_path(base_dir: str, base_name: str) -> str:
    """
    If file exists, add suffix ' (n)'. Returns an available absolute path.
    """
    root, ext = os.path.splitext(base_name)
    candidate = os.path.join(base_dir, f"{root}{ext}")
    n = 1
    while os.path.exists(candidate):
        candidate = os.path.join(base_dir, f"{root} ({n}){ext}")
        n += 1
    return candidate


def write_tasks_xlsx(
    tasks: List[Dict[str, Any]],
    project_name: str,
    out_dir: str,
    extra_headers: Iterable[str] = (),
    project_sheet_rows: Optional[Iterable[Iterable[Any]]] = None,
    template_path: Optional[str] = None,
) -> str:
    """
    Write tasks to an Excel file with a header containing the union of all
    flattened keys. The workbook is based on a template Excel file
    (``excel_template/iquavis_tasks.xlsm`` by default) so that existing
    formatting/VBA can be preserved. A "project" sheet is populated using
    ``project_sheet_rows`` if provided. Returns the absolute path to the written
    file. Provide ``template_path`` to override the template location. When the
    template is macro-enabled (``.xlsm``), VBA content is retained and the
    output file also uses ``.xlsm``. Otherwise, the export falls back to
    ``.xlsx``.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.worksheet.protection import SheetProtection
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to export Excel files. Please install via 'pip install openpyxl'."
        ) from e

    def _sanitize(value: Any) -> Any:
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    # First flatten rows; keep in memory for simplicity and consistent headers
    flat_rows: List[Dict[str, Any]] = [flatten_dict(t) for t in tasks]
    project_rows_list: List[List[Any]] = [list(r) if isinstance(r, (list, tuple)) else list(r or []) for r in project_sheet_rows or []]
    headers = collect_headers(flat_rows, extra_headers=extra_headers)

    template_candidate = template_path or os.path.join(
        EXCEL_TEMPLATE_DIR, DEFAULT_TEMPLATE_NAME
    )
    template_candidate = os.path.abspath(template_candidate)
    if not os.path.exists(template_candidate):
        raise FileNotFoundError(
            "Excel template not found. Place 'iquavis_tasks.xlsm' under the "
            "'excel_template' directory or provide template_path explicitly."
        )

    template_ext = os.path.splitext(template_candidate)[1].lower()
    keep_vba = template_ext == ".xlsm"
    wb = load_workbook(template_candidate, keep_vba=keep_vba)

    def _ensure_sheet(name: str):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(title=name)

    def _clear_values(ws):
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row <= 0 or max_col <= 0:
            return
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.value = None

    ws_project = _ensure_sheet("project")
    if project_rows_list:
        _clear_values(ws_project)
    for row_idx, row in enumerate(project_rows_list, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws_project.cell(row=row_idx, column=col_idx, value=_sanitize(value))

    ws_tasks = _ensure_sheet("tasks")
    _clear_values(ws_tasks)
    for col_idx, header in enumerate(headers, start=1):
        ws_tasks.cell(row=1, column=col_idx, value=_sanitize(header))
    for row_idx, row in enumerate(flat_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws_tasks.cell(row=row_idx, column=col_idx, value=_sanitize(row.get(header)))

    ws_orig = _ensure_sheet("_original")
    _clear_values(ws_orig)
    for col_idx, header in enumerate(headers, start=1):
        ws_orig.cell(row=1, column=col_idx, value=_sanitize(header))
    for row_idx, row in enumerate(flat_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws_orig.cell(row=row_idx, column=col_idx, value=_sanitize(row.get(header)))
    ws_orig.sheet_state = "hidden"

    # Highlight cells edited by users compared to the original sheet
    if headers:
        max_col_letter = get_column_letter(len(headers))
        max_row = len(flat_rows) + 1
        data_range = f"A1:{max_col_letter}{max_row}"
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        ws_tasks.conditional_formatting.add(
            data_range, FormulaRule(formula=["A1<>_original!A1"], fill=yellow_fill)
        )

        # Allow editing cells but disallow row/column manipulation
        for row in ws_tasks.iter_rows(
            min_row=1, max_row=max_row, max_col=len(headers)
        ):
            for cell in row:
                cell.protection = Protection(locked=False)

    ws_tasks.protection = SheetProtection(
        sheet=True,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        deleteColumns=False,
        deleteRows=False,
        sort=False,
        autoFilter=False,
        pivotTables=False,
    )

    safe_name = sanitize_filename(project_name)
    output_ext = ".xlsm" if keep_vba else ".xlsx"
    file_name = f"tasks_{safe_name}{output_ext}"
    out_path = next_available_path(out_dir, file_name)

    wb.save(out_path)
    return out_path
